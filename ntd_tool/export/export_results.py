# ntd_tool/export/export_results.py
from __future__ import annotations

import pandas as pd
from pathlib import Path
from datetime import datetime
from typing import List

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, PatternFill, Font

# ------------------------------------------------------------------ #
#                    INTERNAL CONSTANTS / HELPERS                    #
# ------------------------------------------------------------------ #
KEY_COLS = [
    "title",
    "abstract",
    "doi",
    "pub_year",
    "pub_month",
    "publication_type",
    "country",
    "agency",
    "_source",
    "language",
    "mesh_terms",
]

META_COLS = ["language", "publication_type", "country", "agency"]

INCL_MODELS = ["bert", "svm", "gemini"]  # order for prob / disease cols


def _prep_model_df(df: pd.DataFrame, tag: str) -> pd.DataFrame:
    df = df.copy()

    # Rename the model-specific columns
    if "inclusion_prob" in df.columns:
        df = df.rename(columns={"inclusion_prob": f"inclusion_probability_{tag}"})
    if "disease_predictions" in df.columns:
        df[f"identified_diseases_{tag}"] = df["disease_predictions"].apply(
            lambda x: ", ".join(x) if isinstance(x, (list, tuple, set)) else ""
        )

    wanted: List[str] = [c for c in KEY_COLS if c in df.columns]
    wanted += [
        f"inclusion_probability_{tag}",
        f"identified_diseases_{tag}",
        "mesh_terms",  # ensure present if available
    ]
    # preserve order, dedup
    keep, seen = [], set()
    for c in wanted:
        if c in df.columns and c not in seen:
            keep.append(c)
            seen.add(c)
    return df[keep]


def _reorder_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Return df with columns in workflow-friendly order."""
    desired: List[str] = []

    # 1. Core text
    for col in ["title", "abstract"]:
        if col in df.columns:
            desired.append(col)

    # 2. Inclusion probabilities
    for m in INCL_MODELS:
        col = f"inclusion_probability_{m}"
        if col in df.columns:
            desired.append(col)

    # 3. Paper link
    if "paper_link" in df.columns:
        desired.append("paper_link")

    # 4. Pub year
    if "pub_year" in df.columns:
        desired.append("pub_year")

    # 5. Identified diseases
    for m in INCL_MODELS:
        col = f"identified_diseases_{m}"
        if col in df.columns:
            desired.append(col)

    # 6. Mesh terms
    if "mesh_terms" in df.columns:
        desired.append("mesh_terms")

    # 7. Remaining metadata / identifiers
    for col in META_COLS + ["pub_month", "doi", "_source"]:
        if col in df.columns and col not in desired:
            desired.append(col)

    # 8. Any other stray columns
    for col in df.columns:
        if col not in desired:
            desired.append(col)

    return df[desired]


# ------------------------------------------------------------------ #
#                         MAIN EXPORT FUNCTION                       #
# ------------------------------------------------------------------ #
def export_results(
    svm_df: pd.DataFrame | None = None,
    bert_df: pd.DataFrame | None = None,
    gemini_df: pd.DataFrame | None = None,
    output_path: str | None = None,
) -> str:
    model_dfs: List[pd.DataFrame] = []

    if svm_df is not None:
        model_dfs.append(_prep_model_df(svm_df, "svm"))
    if bert_df is not None:
        model_dfs.append(_prep_model_df(bert_df, "bert"))
    if gemini_df is not None:
        model_dfs.append(_prep_model_df(gemini_df, "gemini"))

    if not model_dfs:
        print("⚠️  No predictions to export.")
        return ""

    # 1. Merge on KEY_COLS subset that actually exists
    merged_df = model_dfs[0]
    for df in model_dfs[1:]:
        merged_df = pd.merge(
            merged_df,
            df,
            on=[c for c in KEY_COLS if c in merged_df.columns],
            how="outer",
        )

    # 2. Build paper link
    def build_link(row):
        if pd.notna(row.get("doi")) and str(row["doi"]).strip():
            return f"https://doi.org/{row['doi']}"
        if pd.notna(row.get("title")):
            return f"https://scholar.google.com/scholar?q={row['title']}"
        return ""

    merged_df["paper_link"] = merged_df.apply(build_link, axis=1)

    # 3. Sort by best probability
    for m in INCL_MODELS:
        col = f"inclusion_probability_{m}"
        if col in merged_df.columns:
            merged_df = merged_df.sort_values(col, ascending=False)
            break

    # 4. Re-order columns
    merged_df = _reorder_columns(merged_df)

    # 5. Write Excel
    export_dir = Path("exports")
    export_dir.mkdir(exist_ok=True)
    if output_path is None:
        output_path = export_dir / f"ntd_results_{datetime.now():%Y-%m-%d_%H-%M}.xlsx"
    else:
        output_path = Path(output_path)

    merged_df.to_excel(output_path, index=False, sheet_name="Results")

    # 6. Format with openpyxl
    wb = load_workbook(output_path)
    ws = wb.active

    # Freeze header row + Title column
    ws.freeze_panes = "B2"  # column A and row 1 stay visible

    # Header style
    header_fill = PatternFill("solid", start_color="B7E1F7", end_color="B7E1F7")
    header_align = Alignment(wrap_text=True, vertical="center", horizontal="center")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.alignment = header_align
        cell.font = Font(bold=True)
    ws.row_dimensions[1].height = 35

    # Column widths and cell wrapping
    specific_widths = {"title": 60, "abstract": 100}
    default_width = 25
    wrap_align = Alignment(wrap_text=True, vertical="top", horizontal="left")

    for idx, column in enumerate(ws.iter_cols(1, ws.max_column), 1):
        header = column[0].value
        ws.column_dimensions[get_column_letter(idx)].width = specific_widths.get(header, default_width)
        for cell in column[1:]:
            cell.alignment = wrap_align
            ws.row_dimensions[cell.row].height = 30  # taller rows

    wb.save(output_path)
    print(f"✅ Exported to {output_path}")
    return str(output_path)
